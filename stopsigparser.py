#Use this code for dev as of 11.9.21

#The Art VanDeLay imports/exports... but more imports than exports... and that's his problem
import pandas as pd
import os
import re
from itertools import chain
from datetime import date
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from collections import Counter
from openpyxl import load_workbook



#These functions work for all file types... note the "specific" functions below
class MainInfoParser():
    
    def __init__(self,file,dates_list,start_time_list,subjects_list,msn_list,day_list):
        self.file = file
        self.dates = dates_list
        self.start = start_time_list
        self.subjects = subjects_list
        self.msns = msn_list
        self.days = day_list
        
    
    def maininfograbber(self):
        with open(self.file,"r") as f:
            datefinder = re.search(r"Start Date:" + ".+" + "\n",f.read())
            dategrabber = datefinder.group().split(":")
            self.dates.append(pd.to_datetime(dategrabber[1].strip()).date())
        with open(self.file,"r") as f:
            subjectfinder = re.search(r"Subject:" + ".+" + "\n", f.read())
            subjectgrabber = subjectfinder.group().split(":")
            self.subjects.append(subjectgrabber[1].strip())
        with open(self.file,"r") as f:
            msnfinder = re.search(r"MSN:" + ".+" + "\n", f.read())
            msngrabber = msnfinder.group().split(":")
            self.msns.append(msngrabber[1].strip())
        with open(self.file,"r") as f:
            paradigmfinder = re.search(r"Experiment:" + ".+" + "\n",f.read())
            paradigmgrabber = paradigmfinder.group().split(": ")
            self.days.append(paradigmgrabber[1].strip())
        with open(self.file,"r") as f:
            startfinder = re.search(r"Start Time:" + ".+" + "\n",f.read())
            startgrabber = startfinder.group().split(": ")
            self.start.append(pd.to_datetime(startgrabber[1].strip()).time())
            
            
    #Moved as static methods w/n class for namespace purposes- both are considered main info, but a small subset
    @staticmethod
    def msngrabber(file):
        with open(file,"r") as f:
            msnfinder = re.search(r"MSN:" + ".+" + "\n", f.read())
            msngrabber = msnfinder.group().split(":")
            return msngrabber[1].strip()
    @staticmethod
    def dayofparadigm(file):
        with open(file,"r") as f:
            paradigmfinder = re.search(r"Experiment:" + ".+" + "\n",f.read())
            paradigmgrabber = paradigmfinder.group().split(": ")
            return paradigmgrabber[1].strip()
    @staticmethod
    def subject_grabber(file):
        with open(file,"r") as f:
            subjectfinder = re.search(r"Subject:" + ".+" + "\n", f.read())
            subjectgrabber = subjectfinder.group().split(":")
            return subjectgrabber[1].strip()

        

class ArrayParser():
    
    def __init__(self,file,array_to_append,upper_delimiter,lower_delimiter):
        self.file = file
        self.array_to_ap = array_to_append
        self.upper = upper_delimiter
        self.lower = lower_delimiter


    def arraygrabber(self):
        with open(self.file,"r") as f:
            arrayfinder = re.search(fr"\n{self.upper}:", f.read())
            arraystart = arrayfinder.start()
        with open(self.file,"r") as f:
            arrayendfinder = re.search(fr"\n{self.lower}:", f.read())
            arrayend = arrayendfinder.start()
        with open(self.file,"r") as f:
            arrayfull = (f.read()[arraystart:arrayend + 1])
            arraystrip = arrayfull.strip(" ")
            arraysplitnewline = arraystrip.split("\n")
            arrayformat1 = (i.strip(" ") for i in arraysplitnewline)
            arrayformatcolumn = [re.split(r"\s{2,8}",i) for i in arrayformat1]
        return self.array_to_ap.append(list(chain.from_iterable(i[1:] for i in arrayformatcolumn)))


    def endarraygrabber(self):
        with open(self.file,'r') as f:
            arrayfinder = re.search(fr'\n{self.upper}:', f.read())
            arraystart = arrayfinder.start()
        with open(self.file,'r') as f:
            arrayfull = (f.read()[arraystart:])
            arraystrip = arrayfull.strip(" ")
            arraysplitnewline = arraystrip.split('\n')
            arrayformat1 = (i.strip(' ') for i in arraysplitnewline)
            arrayformat2 = [re.split(r'\s{2,8}',i) for i in arrayformat1]
            return self.array_to_ap.append(list(chain.from_iterable(i[1:] for i in arrayformat2)))






class Parse():

    def __init__(self):
        self.dates = []
        self.starttimes = []
        self.subjects = []
        self.msns = []
        self.paradigms= []
        self.total_stop_trials = []
        self.stop_correct = []
        self.stop_correct_percent = []
        self.stop_incorrect = []
        self.stop_incorrect_percent = []
        self.go_trial_latencies = []
        self.go_trial_q1 = []
        self.go_trial_q2 = []
        self.go_trial_q3 = []
        self.go_correct = []
        self.delay_lengths_used = []
        self.stop_rxn_times = []
        self.stop_rxn_stds = []
        self.all_stop_rxn_times = []
        self.all_go_latencies = []
        self.all_delays = []
        self.stop_uts = []
        self.stop_sts= []
        self.all_sx_delay_count = []
        self.all_unsx_delay_count = []
        self.file_path = os.getcwd()
        self.sx_test = []
        self.unsx_test = []

        #Grab group identifiers AND grab acceptable msn names
        self.dfgroups = pd.ExcelFile(self.file_path+'\\Group Identifier.xlsx').parse() #FILE PATH!
        self.control_name = self.dfgroups.keys()[0] #Handles group ID name changes
        self.experimental_name = self.dfgroups.keys()[1] #Handles group ID name changes
        self.listofcontrols = [str(i).upper() for i in self.dfgroups[self.control_name]] #Because df to df comparisons weren't working...
        self.listofexps = [str(i).upper() for i in self.dfgroups[self.experimental_name]] #Because df to df comparisons weren't working...
        self.accepted_msns = [str(i).upper().replace(' ','') for i in self.dfgroups['Acceptable MSNs']]

        #The directory where the data files are that you want to analyze
        self.path_to_data = self.file_path +'\\Data\\'
        self.fig_save_loc = self.file_path + '\\Figures\\'
        self.collect()
        
        

    def collect(self):
        #Main loop --> operates outside of def main() so it runs without instantiating python --> stopsigparser.main()
        for i in os.listdir(self.path_to_data):
            self.file = self.path_to_data + i
            msn_check = MainInfoParser.msngrabber(self.file)
            if msn_check.upper().replace(' ','') in self.accepted_msns:
                main_info = MainInfoParser(self.file,self.dates,self.starttimes,self.subjects,self.msns,self.paradigms)
                main_info.maininfograbber()

                
                temp_num_stop_correct = []
                get_stop_correct = ArrayParser(self.file, temp_num_stop_correct, 'H','I')
                get_stop_correct.arraygrabber()
                self.stop_correct.append(int(float(temp_num_stop_correct[-1][0])))
                
                #find the incorrect stop latencies only)
                incorrect_stops_pre = []
                find_incorrect_stops = ArrayParser(self.file,incorrect_stops_pre, 'F','G')
                find_incorrect_stops.arraygrabber()
                incorrect_stops = int(float(incorrect_stops_pre[-1][0]))
                self.stop_incorrect.append(incorrect_stops)
                
                
                self.total_stop_trials.append(self.stop_incorrect[-1]+ self.stop_correct[-1])
                self.stop_incorrect_percent.append(self.stop_incorrect[-1]/self.total_stop_trials[-1]*100)
                
                delay_lengths_pre = []
                find_delays = ArrayParser(self.file, delay_lengths_pre, 'T','V')
                find_delays.arraygrabber()
                #delay_lengths_pre = [float(x) for i,x in enumerate(delay_lengths_pre[-1]) if i != 0 and float(x) > 0.0] #fix if greg fixes program
                self.delay_lengths_used.append(list(set(delay_lengths_pre[-1])))
                
                #You may have to delete this later if Greg fixes the code for the boxes to input correct delay in array T
                temp_q = []
                get_q = ArrayParser(self.file, temp_q, 'Q', 'R')
                get_q.arraygrabber()
                self.q = float(temp_q[-1][0])
                
                
                self.all_delays.append([self.q-float(i) for i in delay_lengths_pre[-1][1:self.total_stop_trials[-1]+1]])
                
                temp_stop_sts = []
                get_stop_sts = ArrayParser(self.file, temp_stop_sts, 'V', 'X')
                get_stop_sts.arraygrabber()
                self.stop_sts.append([float(i) for i in temp_stop_sts[-1][1:self.stop_correct[-1]+1]])
                
                temp_stop_uts = []
                get_stop_uts = ArrayParser(self.file, temp_stop_uts, 'O','T')
                get_stop_uts.arraygrabber()
                self.stop_uts.append([float(i) for i in temp_stop_uts[-1][1:self.stop_incorrect[-1]+1]])
                
                
                #pull all of the stop latencies for assessment of correct vs incorrect
                stop_latencies = []
                find_stop_lat = ArrayParser(self.file, stop_latencies, 'Z', 'NONE')
                find_stop_lat.endarraygrabber()
                stop_latencies = [float(x) for i,x in enumerate(stop_latencies[-1][1:len(delay_lengths_pre)+1])] #fix if greg fixes program
                if stop_latencies.count(0.0) == 0:
                    self.stop_correct_percent.append(0)
                else:
                    self.stop_correct_percent.append(stop_latencies.count(0.0)/len(delay_lengths_pre)*100)
                
                #grab only the stop latencies for reaction time
                rxn_stop_latencies = [x for i,x in enumerate(stop_latencies[:incorrect_stops])]
                self.all_stop_rxn_times.append(rxn_stop_latencies)
                mean_rxn = float(np.mean(rxn_stop_latencies))
                self.stop_rxn_times.append(mean_rxn)
                stop_rxn_std_pre = float(np.std(rxn_stop_latencies))
                self.stop_rxn_stds.append(stop_rxn_std_pre)
                
                    
                    
                total_go_trials = []
                find_go_total = ArrayParser(self.file, total_go_trials, 'G','H')
                find_go_total.arraygrabber()
                total_go_trials = int(float(total_go_trials[-1][0]))
                correct_gos = []
                find_correct_gos = ArrayParser(self.file, correct_gos, 'D','E')
                find_correct_gos.arraygrabber()
                correct_gos = int(float(correct_gos[-1][0]))
                go_trials = []
                find_gos = ArrayParser(self.file, go_trials, 'X','Z')
                find_gos.arraygrabber()
                go_trials =[float(i) for i in go_trials[-1][1:correct_gos+1]] #Fix if greg fixes program
                self.all_go_latencies.append(go_trials)
                self.go_trial_latencies.append(sum(go_trials)/len(go_trials))
                self.go_trial_q1.append(np.percentile(go_trials,25))
                self.go_trial_q2.append(np.percentile(go_trials,50))
                self.go_trial_q3.append(np.percentile(go_trials,75))
                self.go_correct.append(correct_gos/total_go_trials*100)
                
                
                success_id = {'Unsx':self.stop_uts[-1], 'Sx': self.stop_sts[-1]}
                stts = [i for i in self.stop_sts[-1]]
                stts.extend(self.stop_uts[-1])
                stts.sort()

                sx_outcome=[]
                unsx_outcome= []
                sx_outcome.append([self.all_delays[-1][stts.index(i)]  for i in stts if i in success_id['Sx']])
                unsx_outcome.append([self.all_delays[-1][stts.index(i)] for i in stts if i in success_id['Unsx']])

                unsx_outcome = [i for i in unsx_outcome[0]]
                sx_outcome = [i for i in sx_outcome[0]]
#                 unsx_outcome.sort()
#                 sx_outcome.sort()
#                 self.unsx_test.append(unsx_outcome)
#                 self.sx_test.append(sx_outcome)
                self.sx_delay_count = {i:sx_outcome.count(i) for i in list(set(self.all_delays[-1]))}
                self.unsx_delay_count = {i:unsx_outcome.count(i) for i in list(set(self.all_delays[-1]))}
                
                sx_desired_order = list(self.sx_delay_count.keys())
                sx_desired_order.sort()
                self.all_sx_delay_count.append({i:self.sx_delay_count[i] for i in sx_desired_order})

                unsx_desired_order = list(self.unsx_delay_count.keys())
                unsx_desired_order.sort()
                self.all_unsx_delay_count.append({i:self.unsx_delay_count[i] for i in unsx_desired_order})
            else:
                pass

            
        #Create a dictionary that will be used to make the pd.DataFrame    
        df_maker = {'Subject': self.subjects,
                   'Date': self.dates,
                   'Start Time': self.starttimes,
                   'Program': self.msns,
                   'Day of Stop Signal': self.paradigms,
                    'Go Trial % Correct': self.go_correct,
                    'Avg. Go Trial Latency (secs)': self.go_trial_latencies,
                    'Go Trial Latency Q1 (secs)':self.go_trial_q1,
                    'Go Trial Latency Q2 (secs)(median)': self.go_trial_q2,
                    'Go Trial Latency Q3 (secs)': self.go_trial_q3,
                    'Stop Sig Delay Length': self.delay_lengths_used,
                    'Total # Stop Trials': self.total_stop_trials,
                    '# Correct Stop Trials': self.stop_correct,
                    '# Incorrect Stop Trials': self.stop_incorrect,
                   'Stop Sig % Correct': self.stop_correct_percent,
                    'Stop Sig % Incorrect': self.stop_incorrect_percent,
                   'Stop Sig Rxn Times (secs)': self.stop_rxn_times,
                   'Stop Sig Rxn Std': self.stop_rxn_stds,
                   'All Stop Rxn Times': self.all_stop_rxn_times,
                   'All Go Latencies': self.all_go_latencies,
                   'Unsuccessful Counts x Delay Time': self.all_unsx_delay_count,
                   'Successful Counts x Delay Time': self.all_sx_delay_count}
        

                

        #Create dataframe and sort the values by subject/date
        self.df = pd.DataFrame(df_maker)
        self.df.sort_values(['Subject', 'Date'], ascending = (True, True), inplace = True)


        #Create a 'day number' column by animal (i.e., see what day of the paradigm each individual animal is on)
        range_by_animal = [] #This is a list for collecting all the day numbers- needs to be after the sort
        for i in self.df.groupby('Subject'):
            x = range(1,len(i[1])+1)
            for num in x:
                range_by_animal.append(num)
        self.df.insert(1,'Day Number', range_by_animal)


        #Code for assigning a group type to each animal --> uses dfgroup that was created before the main loop
        group_column = []
        for i in self.df['Subject']:
            if i.upper() in self.listofcontrols:
                group_column.append(self.dfgroups.columns[0]) #You can change the names of the columns to match the study!
            elif i.upper() in self.listofexps:
                group_column.append(self.dfgroups.columns[1]) #You can change the names of the columns to match the study!
            else:
                group_column.append('NaN') #Because we need to match the df lengths
                print(f'{i} is not in your Group Identifier spreadsheet!!!! Please Check!!!')
        self.df.insert(0,'Group', group_column)


        #Set the index to be the subject, elimating the autogenerated DataFrame index
        self.df.set_index('Subject', inplace=True)


        #Create a file save path and save a sheet/animal in a workbook for easy data visualization
        data_save = self.file_path + f'\\XL Files\\Stop Signal Data from {date.today()}.xlsx'
        with pd.ExcelWriter(data_save, engine = 'openpyxl') as writer:
            for i,x in self.df.groupby('Subject'):
                x.to_excel(writer, sheet_name = i)
            self.df.groupby('Group').mean().to_excel(writer, sheet_name = 'GROUP AVERAGES')
            self.df.groupby('Group').sem().to_excel(writer, sheet_name = 'GROUP SEM')
            self.df.groupby(['Group','Day Number']).mean().to_excel(writer, sheet_name = 'GROUP X DAY AVERAGES')
            self.df.groupby(['Group','Day Number']).sem().to_excel(writer, sheet_name = 'GROUP X DAY SEM')
            
                #THIS IS STILL IN DEVELOPMENT!!!   
    def gr_rxns(self):
        #Create your iterator for your going through your data and making a graph for each animal that is present
        set_subs = list(set(self.subjects))
        x = round(np.sqrt(len(set_subs)))
        y = round(np.sqrt(len(set_subs)))

        #Create a dataframe for each animal that is present and explode their rxn times
        df_rxn = self.df.explode('All Stop Rxn Times')
        df_rxn['All Stop Rxn Times'] = df_rxn['All Stop Rxn Times'].astype('float')

        #Create your rxn graphs    
        fig,axes = plt.subplots(x,y, figsize = (24,36))
        a= 0
        b= 0
        for i in set_subs:
            sns.violinplot(x = 'Day Number', y = 'All Stop Rxn Times', data = df_rxn[df_rxn.index == i], ax = axes[a,b], palette = 'nipy_spectral')
            axes[a,b].set_title(i)
            if b >= y-1:
                a += 1
                b = 0
            else:
                b += 1

        plt.savefig(self.fig_save_loc+ f'Stop Signal Rxn Times {self.control_name} vs. {self.experimental_name} {date.today()}.png', dpi = 300)

    def gr_go(self):
        set_subs = list(set(self.subjects))
        x = round(np.sqrt(len(set_subs)))
        y = round(np.sqrt(len(set_subs)))
        #Create all of the go latency violin plots!    
        fig,axes = plt.subplots(x,y, figsize = (24,36))
        df_go = self.df.explode('All Go Latencies')
        df_go['Go Latencies'] = df_go['All Go Latencies'].astype('float')

        a= 0
        b= 0
        for i in set_subs:
            sns.violinplot(x = 'Day Number', y = 'Go Latencies', data = df_go[df_go.index == i], ax = axes[a,b], palette = 'nipy_spectral')
            axes[a,b].set_title(i)
            if b >= y-1:
                a += 1
                b = 0
            else:
                b += 1
        plt.savefig(self.fig_save_loc+ f'Go Trial Latencies {self.control_name} vs. {self.experimental_name} {date.today()}.png', dpi = 300)


        
        
        
        
        
            
class SS_Breakdown():

    def __init__(self):
        
        self.file_path = os.getcwd()

        #Grab group identifiers AND grab acceptable msn names
        self.dfgroups = pd.ExcelFile(self.file_path+'\\Group Identifier.xlsx').parse() #FILE PATH!
        self.control_name = self.dfgroups.keys()[0] #Handles group ID name changes
        self.experimental_name = self.dfgroups.keys()[1] #Handles group ID name changes
        self.listofcontrols = [str(i).upper() for i in self.dfgroups[self.control_name]] #Because df to df comparisons weren't working...
        self.listofexps = [str(i).upper() for i in self.dfgroups[self.experimental_name]] #Because df to df comparisons weren't working...
        self.accepted_msns = [str(i).upper().replace(' ','') for i in self.dfgroups['Acceptable MSNs']]

        #The directory where the data files are that you want to analyze
        self.path_to_data = self.file_path +'\\Data\\'
        self.fig_save_loc = self.file_path + '\\Figures\\'
        self.data_save_loc = self.file_path +'\\XL Files\\'
        
        
        
        
        self.breakdown()
        
        
    def send_it(self):
        data_save = self.data_save_loc + f'Stop Signal Data from {date.today()}.xlsx'
        book = load_workbook(data_save)
        sn = self.df.index[0]
        with pd.ExcelWriter(data_save, engine ='openpyxl') as writer:
            writer.book = book
            writer.sheets = {worksheet.title:worksheet for worksheet in book.worksheets}
            self.df.to_excel(writer, startrow= writer.book[sn].max_row+2, sheet_name = sn)     


    def breakdown(self):
        ss_counter = 0
        self.set_of_subs = []
        #Main loop --> operates outside of def main() so it runs without instantiating python --> stopsigparser.main()
        for i in os.listdir(self.path_to_data):
            self.file = self.path_to_data + i
            msn_check = MainInfoParser.msngrabber(self.file)
            if msn_check.upper().replace(' ','') in self.accepted_msns:
                self.set_of_subs.append(MainInfoParser.subject_grabber(self.file))
                
        self.set_of_subs = list(set(self.set_of_subs))
        
        for i in self.set_of_subs:
            self.dates = []
            self.starttimes = []
            self.subjects = []
            self.msns = []
            self.paradigms= []
            self.total_stop_trials = []
            self.stop_correct = []
            self.stop_correct_percent = []
            self.stop_incorrect = []
            self.stop_incorrect_prob = []

            self.num_go_trials = []
            self.delay_lengths_used = []
            self.stop_rxn_times = []
            self.stop_rxn_stds = []
            self.all_stop_rxn_times = []
            self.all_go_latencies = []
            self.all_delays = []
            self.stop_uts = []
            self.stop_sts= []
            self.all_sx_delay_count = []
            self.all_unsx_delay_count = []
            self.correct_gos = []
            self.gts = []
            self.get_ssrt = []
            
            
            
            for x in os.listdir(self.path_to_data):
                
                if i in x:
                    self.file = self.path_to_data + x
                    msn_check = MainInfoParser.msngrabber(self.file)
                    if msn_check.upper().replace(' ','') in self.accepted_msns:
                        main_info = MainInfoParser(self.file,self.dates,self.starttimes,self.subjects,self.msns,self.paradigms)
                        main_info.maininfograbber()
                        
                        
                        #####----------------------------------------------------------------------##############
                        #####EVERYTHING BELOW THIS LINE UNTIL THE NOTIFIER IS FOR THE DEVELOPMENT OF THE STOP SIG
                        #####INFORMATION.THIS SHOULD NOT BE MOVED.###############################################
                        temp_num_stop_correct = []
                        get_stop_correct = ArrayParser(self.file, temp_num_stop_correct, 'H','I')
                        get_stop_correct.arraygrabber()
                        self.stop_correct.append(int(float(temp_num_stop_correct[-1][0])))

                        #find the incorrect stop latencies only)
                        incorrect_stops_pre = []
                        find_incorrect_stops = ArrayParser(self.file,incorrect_stops_pre, 'F','G')
                        find_incorrect_stops.arraygrabber()
                        incorrect_stops = int(float(incorrect_stops_pre[-1][0]))
                        self.stop_incorrect.append(incorrect_stops)

                        self.total_stop_trials.append(self.stop_incorrect[-1]+ self.stop_correct[-1])
                        self.stop_incorrect_prob.append(self.stop_incorrect[-1]/self.total_stop_trials[-1])

                        delay_lengths_pre = []
                        find_delays = ArrayParser(self.file, delay_lengths_pre, 'T','V')
                        find_delays.arraygrabber()
                        #delay_lengths_pre = [float(x) for i,x in enumerate(delay_lengths_pre[-1]) if i != 0 and float(x) > 0.0] #fix if greg fixes program
                        self.delay_lengths_used.append(list(set(delay_lengths_pre[-1])))

                        #You may have to delete this later if Greg fixes the code for the boxes to input correct delay in array T
                        temp_q = []
                        get_q = ArrayParser(self.file, temp_q, 'Q', 'R')
                        get_q.arraygrabber()
                        self.q = float(temp_q[-1][0])


                        self.all_delays.append([(self.q-float(i))/1000 for i in delay_lengths_pre[-1][1:self.total_stop_trials[-1]+1]])

                        temp_stop_sts = []
                        get_stop_sts = ArrayParser(self.file, temp_stop_sts, 'V', 'X')
                        get_stop_sts.arraygrabber()
                        self.stop_sts.append([float(i) for i in temp_stop_sts[-1][1:self.stop_correct[-1]+1]])


                        temp_stop_uts = []
                        get_stop_uts = ArrayParser(self.file, temp_stop_uts, 'O','T')
                        get_stop_uts.arraygrabber()
                        self.stop_uts.append([float(i) for i in temp_stop_uts[-1][1:self.stop_incorrect[-1]+1]])

                        success_id = {'Unsx':self.stop_uts[-1], 'Sx': self.stop_sts[-1]}
                        stts = [i for i in self.stop_sts[-1]]
                        stts.extend(self.stop_uts[-1])
                        stts.sort()

                        sx_outcome=[]
                        unsx_outcome= []
                        sx_outcome.append([self.all_delays[-1][stts.index(i)]  for i in stts if i in success_id['Sx']])
                        unsx_outcome.append([self.all_delays[-1][stts.index(i)] for i in stts if i in success_id['Unsx']])
                        unsx_outcome = [i for i in unsx_outcome[0]]
                        sx_outcome = [i for i in sx_outcome[0]]

                        self.sx_delay_count = {i:sx_outcome.count(i) for i in list(set(self.all_delays[-1]))}

                        self.unsx_delay_count = {i:unsx_outcome.count(i) for i in list(set(self.all_delays[-1]))}

                        sx_desired_order = list(self.sx_delay_count.keys())
                        sx_desired_order.sort()
                        self.all_sx_delay_count.append({i:self.sx_delay_count[i] for i in sx_desired_order})


                        unsx_desired_order = list(self.unsx_delay_count.keys())
                        unsx_desired_order.sort()
                        self.all_unsx_delay_count.append({i:self.unsx_delay_count[i] for i in unsx_desired_order})
                        ###############---------------------------------------------------------------------#########################
                        ###############--------------------END STOP SIGNAL GATHERING INFORMATION------------#########################
            
                        temp_num_gos = []
                        get_gos = ArrayParser(self.file, temp_num_gos, 'G','H')
                        get_gos.arraygrabber()
                        self.num_go_trials.append(int(float(temp_num_gos[-1][0])))
                        
                        
                        
                        temp_correct_gos = []
                        get_correct_gos = ArrayParser(self.file, temp_correct_gos, 'D', 'E')
                        get_correct_gos.arraygrabber()
                        self.correct_gos.append(int(float(temp_correct_gos[-1][0])))
                        
                        temp_gts = []
                        get_gts = ArrayParser(self.file, temp_gts, 'X', 'Z')
                        get_gts.arraygrabber()
                        self.gts.append([float(i) for i in temp_gts[-1][1:self.correct_gos[-1]+1]])
                        self.gts[-1].sort()
                        
                        self.magic_go = int(np.round(self.correct_gos[-1]*self.stop_incorrect_prob[-1]))
                        if self.magic_go == 0.0:
                            self.get_ssrt.append(self.gts[-1][0])
                        else:
                            self.get_ssrt.append(self.gts[-1][self.magic_go-1])
            
            
            ####THIS CREATES JUST ENOUGH COLUMNS FOR EACH ANIMAL#######
            self.unsx_column = []
            for i in self.all_unsx_delay_count:
                for x in i:
                    if x not in self.unsx_column:
                        self.unsx_column.append(x)
            self.sx_column = []
            for i in self.all_sx_delay_count:
                for x in i:
                    if x not in self.sx_column:
                        self.sx_column.append(x)
            
        
            
            
            dfm = {'Subject': self.subjects,
                  'Date': self.dates,
                  'MSN': self.msns,
                  'Total Go Trials': self.num_go_trials,
                  'Stop incorrect prob': self.stop_incorrect_prob,
                  'Pre-Delay SSRT':self.get_ssrt}
            
            self.df =  pd.DataFrame(dfm)
            self.df['Go RXN Time of interest'] = self.df['Total Go Trials']*self.df['Stop incorrect prob']
            
            for i in self.unsx_column:
                self.df[f'Unsx stop @ {i}'] = [x[i] if i in x else np.nan for x in self.all_unsx_delay_count]
                self.df[f'Sx stop @ {i}'] = [x[i] if i in x else np.nan for x in self.all_sx_delay_count]
                self.df[f'Percent correct @ delay {i}'] = self.df[f'Sx stop @ {i}']/(self.df[f'Sx stop @ {i}'] + self.df[f'Unsx stop @ {i}'])*100
                self.df[f'Percent incorrect @ delay {i}'] = self.df[f'Unsx stop @ {i}']/(self.df[f'Sx stop @ {i}'] + self.df[f'Unsx stop @ {i}'])*100
                
                
            #Create a 'day number' column by animal (i.e., see what day of the paradigm each individual animal is on)
            range_by_animal = [] #This is a list for collecting all the day numbers- needs to be after the sort
            for i in self.df.groupby('Subject'):
                x = range(1,len(i[1])+1)
                for num in x:
                    range_by_animal.append(num)
            self.df.insert(1,'Day Number', range_by_animal)


            #Code for assigning a group type to each animal --> uses dfgroup that was created before the main loop
            group_column = []
            for i in self.df['Subject']:
                if i.upper() in self.listofcontrols:
                    group_column.append(self.dfgroups.columns[0]) #You can change the names of the columns to match the study!
                elif i.upper() in self.listofexps:
                    group_column.append(self.dfgroups.columns[1]) #You can change the names of the columns to match the study!
                else:
                    group_column.append('NaN') #Because we need to match the df lengths
                    print(f'{i} is not in your Group Identifier spreadsheet!!!! Please Check!!!')
            self.df.insert(0,'Group', group_column)


            #Set the index to be the subject, elimating the autogenerated DataFrame index
            self.df.set_index('Subject', inplace=True)
            
            
            #Send the dfs to the spreadsheet that was already made
            self.send_it()

            
            
            
            
def get_data_only():
    get_data = Parse()
    get_ss_breakdown = SS_Breakdown()
    

def graphs_and_data():
    get_data = Parse()
    get_ss_breakdown = SS_Breakdown()
    get_data.gr_rxns()
    get_data.gr_go()
